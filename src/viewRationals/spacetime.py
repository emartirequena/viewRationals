import os
from multiprocessing import Pool, cpu_count, managers
import json
import gc
from openpyxl import Workbook

from rationals import Rational, c
from timing import timing, get_last_duration
from utils import collect, divisors
from config import Config


spacetime = None


def eq_digits(q: str, r: str):
	l = len(q)
	for i in range(l):
		if q == r[i:] + r[:i-l]:
			return True
	return False


class HashRationalsItem:
	def __init__(self, min, max):
		self.min = min
		self.max = max
		self.rationals = []
		self.indexes = {}

	def __del__(self):
		del self.rationals
		del self.indexes

	def add(self, m, reminders, digits, time):
		if not self.min <= m <= self.max:
			return False
		
		if m not in self.indexes:
			self.rationals.append({
				'm': [m],
				'digits': digits,
				'count': 1,
				'time': time
			})
			i = len(self.rationals) - 1
			for r in reminders:
				self.indexes[r] = i
		elif m not in self.rationals[self.indexes[m]]['m']:
			self.rationals[self.indexes[m]]['m'].append(m)
			self.rationals[self.indexes[m]]['count'] += 1

		return True


class HashRationals:
	def __init__(self, num, size=100000):
		self.hash: list[HashRationalsItem] = []
		for i in range(0, num, size):
			self.hash.append(HashRationalsItem(i, i+size-1))

	def __del__(self):
		del self.hash

	def add(self, m, reminders, digits, time):
		for item in self.hash:
			if item.add(m, reminders, digits, time):
				return True
	
	def get_rationals(self):
		rationals = []
		for item in self.hash:
			rationals += item.rationals
		return rationals


class Cell(object):
	def __init__(self, dim, T, n, x, y=0, z=0):
		self.dim = dim
		self.T = T
		self.n = n
		self.x = x
		self.y = y
		self.z = z
		self.count = 0
		self.time = 0.0
		self.next_digits = dict(zip([x for x in range(2**self.dim)], [0 for _ in range(2**self.dim)]))
		self.rationals = HashRationals(self.n)

	def __del__(self):
		del self.next_digits
		del self.rationals

	def add(self, time: int, reminders: list[int], digits: str, m: int, next_digit: int):
		self.count += 1
		self.time += time
		self.next_digits[next_digit] += 1
		self.rationals.add(m, reminders, digits, time)

	def clear(self):
		self.count = 0
		self.time = 0.0
		self.next_digits = dict(zip([x for x in range(2**self.dim)], [0 for _ in range(2**self.dim)]))
		self.rationals = HashRationals(self.n)

	def get(self):
		pos = (self.x, )
		if self.dim > 1:
			pos = pos + (self.y, )
		if self.dim > 2:
			pos = pos + (self.z, )
		out = {
			'pos': pos,
			'count': self.count,
			'time': self.time / float(self.count),
			'next_digits': self.next_digits,
			'rationals': self.rationals.get_rationals()
		}
		return out
	
	def set(self, count, time, next_digits, rationals):
		self.count = count
		self.time = time
		self.next_digits = dict(zip([x for x in range(2**self.dim)], [0 for _ in range(2**self.dim)]))
		for x in [x for x in range(2**self.dim)]:
			self.next_digits[x] = next_digits[str(x)]
		self.rationals = HashRationals(self.n)
		for rational in rationals:
			for m in rational['m']:
				self.rationals.add(m, rational['m'], rational['digits'], rational['time'])


class Space(object):
	def __init__(self, t, dim, T, n, name='normal'):
		self.t = t
		self.T = T
		self.n = n
		self.dim = dim
		self.name = name
		self.base = 2**dim
		self.indexes: list[int] = []
		self.cells: list[Cell] = []
		num = (t + 1)**self.dim
		self.indexes = [-1 for _ in range(num)]

	def __del__(self):
		del self.cells

	def getCell(self, x, y=0, z=0):
		nx =  c * self.t - x
		ny = (c * self.t - y) if self.dim > 1 else 0.0
		nz = (c * self.t - z) if self.dim > 2 else 0.0
		n = int(nx + (self.t + 1) * (ny + (self.t + 1) * nz))
		if n < 0 or n >= len(self.indexes):
			return None
		if self.indexes[n] < 0:
			self.indexes[n] = len(self.cells)
			self.cells.append(Cell(self.dim, self.T, self.n, x, y, z))
		return self.cells[self.indexes[n]]
	
	def countCells(self):
		return len(self.cells)

	def getCells(self):
		return self.cells

	def add(self, time, reminders, digits, m, next_digit, x, y, z):
		cell = self.getCell(x, y, z)
		if not cell:
			return
		cell.add(time, reminders, digits, m, next_digit)

	def clear(self):
		for cell in self.cells:
			del cell
		del self.cells
		self.cells = []
		for n in range(len(self.indexes)):
			self.indexes[n] = -1

	def save(self):
		out_cells = []
		for cell in self.cells:
			out_cells.append(cell.get())
		return out_cells
	
	def load(self, input: list[dict]):
		self.clear()
		for in_cell in input:
			cell = self.getCell(*in_cell['pos'])
			cell.set(in_cell['count'], in_cell['time'], in_cell['next_digits'], in_cell['rationals'])

	def getMaxTime(self):
		max_time = -1
		for cell in self.cells:
			if cell.time > max_time:
				max_time = cell.time
		return max_time


class Spaces:
	def __init__(self, T, n, max, dim=1) -> None:
		self.T = T
		self.n = n
		self.max = max
		self.dim = dim
		self.spaces = [Space(t, dim, T, n) for t in range(max + 1)]
		self.accumulates_even = Space(max if T%2 == 0 else max-1, dim, T, n, name='even')
		self.accumulates_odd  = Space(max if T%2 == 1 else max-1, dim, T, n, name='odd' )

	def __del__(self):
		del self.spaces
		del self.accumulates_even
		del self.accumulates_odd

	def add(self, is_special, t, reminders, digits, m, next_digit, time, cycle, x, y, z):
		self.spaces[t].add(time, reminders, digits, m, next_digit, x, y, z)
		if t < self.max - cycle and is_special:
			return
		if self.dim == 1:
			if (x == t * c or x == -t * c) and is_special:
				return
		elif self.dim == 2:
			if (x == y == t * c or x == y == -t * c) and is_special:
				return
		else:
			if (x == y == z == t * c or x == y == z == -t * c) and is_special:
				return
		if t%2 == 0:
			self.accumulates_even.add(time, reminders, digits, m, next_digit, x, y, z)
		else:
			self.accumulates_odd.add(time, reminders, digits, m, next_digit, x, y, z)

	def getMaxTime(self, accumulate):
		max_time = -1
		if not accumulate:
			for space in self.spaces:
				spc_time = space.getMaxTime()
				if spc_time > max_time:
					max_time = spc_time
			return max_time
		else:
			max_time = self.accumulates_even.getMaxTime()
			odd_max_time = self.accumulates_odd.getMaxTime()
			if odd_max_time > max_time:
				max_time = odd_max_time
		return max_time

	def clear(self):
		for space in self.spaces:
			space.clear()
		self.accumulates_even.clear()
		self.accumulates_odd.clear()

	def getCell(self, t, x, y=0, z=0, accumulate=False):
		if not accumulate:
			return self.spaces[t].getCell(x, y, z)
		else:
			if t%2 == 0:
				return self.accumulates_even.getCell(x, y, z)
			else:
				return self.accumulates_odd.getCell(x, y, z)
			
	def getCells(self, t, accumulate=False):
		if not accumulate:
			return self.spaces[t].getCells()
		else:
			if t%2 == 0:
				return self.accumulates_even.getCells()
			else:
				return self.accumulates_odd.getCells()

	def getSpace(self, t, accumulate=False):
		if not accumulate:
			return self.spaces[t]
		else:
			if t%2 == 0:
				return self.accumulates_even
			else:
				return self.accumulates_odd

	def save(self):
		output = {}
		for t in range(self.max + 1):
			space = self.getSpace(t, accumulate=False)
			out_cells = space.save()
			output[str(t)] = out_cells
		output['accumulates_even'] = self.accumulates_even.save()
		output['accumulates_odd'] = self.accumulates_odd.save()
		return output
	
	def load(self, input: dict):
		for t in range(self.max + 1):
			space = self.getSpace(t)
			space.load(input[str(t)])
		self.accumulates_even.load(input['accumulates_even'])
		self.accumulates_odd.load(input['accumulates_odd'])

class MyManager(managers.BaseManager):
	...

MyManager.register('Spaces', Spaces)

def create_rational(args):
	m, n, dim = args
	return Rational(m, n, dim)


def add_rational1(args):
	_, rt, t, x, y, z = args
	r: Rational = args[0]
	px, py, pz = r.position(rt)
	px += x
	py += y
	pz += z
	reminders = r.reminders
	digits = r.path()
	m = r.m
	next_digit = r.digit(t+rt+1)
	time = r.time(t+rt)
	obj = (t+rt, reminders, digits, m, next_digit, time, px, py, pz)
	return obj


def add_rational2(args):
	spaces, is_special, pT, max, r, t, x, y, z = args
	for rt in range(0, max + 1):
		px, py, pz = r.position(rt)
		px += x
		py += y
		pz += z
		reminders = r.reminders
		digits = r.path()
		m = r.m
		next_digit = r.digit(t+rt+1)
		time = r.time(t+rt)
		spaces.add(is_special, t+rt, reminders, digits, m, next_digit, time, pT, px, py, pz)


class SpaceTime(object):
	def __init__(self, T, n, max, dim=1):
		self.T = T
		self.max = max
		self.dim = dim
		self.n = n
		self.is_special = False
		self.manager = MyManager()
		self.manager.start()
		self.spaces = self.manager.Spaces(T, n, max, dim)
		self.rationalSet = []
		self.algorithm = 2
		self.changed = False

	def __del__(self):
		del self.spaces
		if self.rationalSet:
			del self.rationalSet
		collect()

	def getParams(self):
		return self.T, self.n, self.max, self.dim, self.is_special

	def len(self):
		return self.max

	def clear(self):
		self.n = 0
		self.is_special = False
		self.spaces.clear()
		collect()

	def getCell(self, t, x, y=0, z=0, accumulate=False):
		return self.spaces.getCell(t, x, y, z, accumulate)

	def getCells(self, t, accumulate=False):
		return self.spaces.getCells(t, accumulate)
	
	def getSpace(self, t, accumulate=False):
		return self.spaces.getSpace(t, accumulate)
	
	def getMaxTime(self, accumulate=False):
		return self.spaces.getMaxTime(accumulate)
	
	@timing
	def setRationalSet(self, n: int, is_special: bool = False):
		self.n = n
		self.is_special = is_special
		p = Pool(cpu_count())
		params = []
		for m in range(n + 1):
			params.append((m, n, self.dim))

		unordered_set = p.imap(func=create_rational, iterable=params, chunksize=10000)
		p.close()
		p.join()

		self.rationalSet = list(sorted(unordered_set, key=lambda x: x.m))
		del unordered_set
		collect()

	def set_algorithm(self, algo):
		self.algorithm = algo

	@timing
	def addRationalSet(self, t=0, x=0, y=0, z=0):
		self.spaces.clear()
		print(f'algorithm: {self.algorithm}')

		if self.algorithm == 0:
			for r in self.rationalSet:
				add_rational2((self.spaces, self.is_special, self.T, self.max, r, t, x, y, z))

		elif self.algorithm == 1:
			num_cpus = int(cpu_count() * 0.8)
			chunksize = ((self.max * len(self.rationalSet)) // num_cpus) or 1
			p = Pool(num_cpus)
			params = []
			for r in self.rationalSet:
				for rt in range(self.max + 1):
					params.append((r, rt, t, x, y, z))
			results = p.imap(func=add_rational1, iterable=params, chunksize=chunksize)
			p.close()
			p.join()

			for result in results:
				pt, reminders, digits, m, next_digit, time, px, py, pz = result
				self.spaces.add(self.is_special, pt, reminders, digits, m, next_digit, time, self.T, px, py, pz)
				del result

			del params
			del results
			collect()

		elif self.algorithm == 2:
			num_cpus = int(cpu_count() * 0.8)
			chunksize = ((len(self.rationalSet)) // num_cpus) or 1
			p = Pool(num_cpus)
			params = []
			for r in self.rationalSet:
				params.append((self.spaces, self.is_special, self.T, self.max, r, t, x, y, z))

			p.imap(func=add_rational2, iterable=params, chunksize=chunksize)
			p.close()
			p.join()

			del params
			collect()

		self.changed = False

	def reset(self, T, num, max, dim):
		self.__init__(T, num, max, dim)
		self.changed = True

	@timing
	def save(self, fname):
		spaces = self.spaces.save()
		output = {
			'dim': self.dim,
			'num': self.n,
			'special': self.is_special,
			'T': self.T,
			'max': self.max,
			'spaces': spaces
		}
		with open(fname, 'wt') as fp:
			json.dump(output, fp, indent=4)

	@timing
	def load(self, fname):
		with open(fname, 'rt') as fp:
			input = json.load(fp)
		self.reset(input['T'], input['num'], input['max'], input['dim'])
		self.is_special = input['special']
		self.spaces.load(input['spaces'])


if __name__ == '__main__':
	config = Config()
	path = config.get('files_path')
	fname = os.path.join(path, 'algorithms.xlsx')
	dim = 3
	T = 12
	n = (2**dim)**int(T // 2) + 1
	divisors = divisors(n)
	print('Creating spacetime...')
	spacetime = SpaceTime(T, n, T, dim=dim)

	wb = Workbook()
	ws = wb.active

	ws.cell(row=1, column=1, value='number')
	ws.cell(row=1, column=2, value='algo 0')
	ws.cell(row=1, column=3, value='algo 1')
	ws.cell(row=1, column=4, value='algo 2')

	for algorithm in range(3):
		spacetime.set_algorithm(algorithm)
		row = 2
		for n in divisors:
			ws.cell(row=row, column=1, value=n)

			print(f'Set rational set for n={n}...')
			spacetime.setRationalSet(n, is_special=True)
			print('Add rational set...')
			spacetime.addRationalSet()

			duration = get_last_duration()
			ws.cell(row=row, column=algorithm+2, value=duration)
			row += 1

	wb.save(fname)
		

