import os
import sys
from openpyxl import Workbook
from time import time
from rationals import Rational
from transform import Transform
from multiprocessing import Pool, cpu_count, managers, freeze_support
import gc


def count_ones(seq: list[list[str]]) -> tuple[int]:
    nx = 0
    ny = 0
    nz = 0
    for elem in seq:
        d = int(elem[0])
        nx += d % 2
        ny += (d // 2) % 2
        nz += (d // 4) % 2
    return nx, ny, nz


# class MyManager(managers.BaseManager):
# 	...

# MyManager.register('Transform', Transform)


def transform_digits(args):
    tr, s, mode, form = args
    seqs = tr.transform(s, mode, form)
    return seqs


def render(ws, row, dim, t, n, tr: Transform, mode, form, results: list, outs: dict):
    init = time()

    print(f'preparing transform...')
    outputs = {}
    num_cpus = int(cpu_count() * 0.8)
    chunksize = 100
    p = Pool(num_cpus)
    params = []
    r = Rational(1, n, dim)
    for m in range(n+1):
        r.reset(m, n, dim)
        s = r.path(t)
        params.append((tr, s, mode, form))
    gc.collect()
    print('transform...')
    seqs = p.map(func=transform_digits, iterable=params, chunksize=chunksize)
    p.close()
    p.join()

    print(f'gathering transformed seqs...')
    count = 0
    for l in seqs:
        for seq in l:
            if seq not in outputs:
                outputs[seq] = 0
            outputs[seq] += 1
            count += 1
    print(f'gathered transform of {count} seqs...')

    print('dump...')
    space = {}
    total = 0
    for digits, num in outputs.items():
        nx = 0
        ny = 0
        for d in digits[:t]:
            i = int(d)
            nx += i % 2
            if dim > 1:
                ny += (i // 2) % 2
        pos = nx + (t+1) * ny
        if pos not in space:
            space[pos] = 0
        space[pos] += num
        total += num

    print('render...')
    if dim == 1:
        result = []
        for x in range(t+1):
            pos = x
            num =space.get(pos, 0)
            weight = num / total
            ws.cell(row=row, column=x+1).value = weight
            result.append(weight)
        row += 3

    elif dim > 1:
        result = []
        for y in range(t+1):
            for x in range(t+1):
                pos = x + y * (t+1)
                num = space.get(pos, 0)
                weight = num / total
                ws.cell(row=row, column=x + 1).value = weight
                result.append(weight)
            row += 1
        row += 3

    if result not in results:
        results.append(result)
    index = results.index(result)
    if index not in outs:
        outs[index] = []
    outs[index].append(tr.get_output_form(mode, form))

    end = time()
    tmp = end-init
    min = int(tmp / 60)
    sec = int(tmp) % 60
    print(f'render time: {min:02d}:{sec:02d}')

    return row


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = 'main'

    dim = 2
    n = 205
    nd = 12
    size_trans = 10
    nx = 4

    ws.cell(1, 1).value = n
    ws.cell(2, 1).value = nd
    ws.cell(3, 1).value = size_trans

    row = 5
    for ny in range(size_trans-2, size_trans-1):

        tr = Transform(dim, size_trans, nx, ny)
        print(f'({nx}, {ny}) modes: {tr.get_num_modes()}')

        for mode in range(tr.get_num_modes()):
            ws.cell(row, 1).value = nx
            ws.cell(row, 2).value = ny
            row += 1

            ones = count_ones(tr.output_modes[mode].values())
            ws.cell(row, 1).value = f'mode: {mode}, ones: {ones}'
            row += 1

            print(ny, mode, ones)
            print(tr.plug_input)
            print(tr.output_modes[mode])

            ws.cell(row, 1).value = str(tr.output_modes[mode])
            row += 1

            row = render(ws, row, dim, nd, n, tr, mode)

    wb.save('2d.12b.xlsx')

def main2():
    wb = Workbook()
    ws = wb.active
    ws.title = 'main'

    dim = 2
    n = 4**8-1
    nd = 8
    size_trans = 8
    nx = 3
    ny = 4

    ws.cell(1, 1).value = n
    ws.cell(2, 1).value = nd
    ws.cell(3, 1).value = size_trans

    row = 5
    for t in range(0, nd+1):

        tr = Transform(dim, size_trans, nx, ny)
        print(f'modes: {tr.get_num_modes()}')

        for mode in range(tr.get_num_modes()):
            ws.cell(row, 1).value = nx
            ws.cell(row, 2).value = ny
            row += 1

            ones = count_ones(tr.output_modes[mode].values())
            ws.cell(row, 1).value = f'mode: {mode}, ones: {ones}'
            row += 1

            print(ny, mode, ones)
            print(tr.plug_input)
            print(tr.output_modes[mode])

            ws.cell(row, 1).value = str(tr.output_modes[mode])
            row += 1

            row = render(ws, row, dim, t, n, tr, mode)

    fname = '2d.13.xlsx'
    print(f'escribiendo {fname}')
    wb.save(fname)


def main3(fname, dim, n, T, nd, nx, ny=0, nz=0):
    print('main3', fname, dim, n, T, nd, nx, ny, nz)
    wb = Workbook()
    wb.remove(wb.active)
    # manager = MyManager()
    # manager.start()
    # tr = manager.Transform(dim, nd, nx, ny)
    tr = Transform(dim, nd, nx, ny, nz)

    print(f'({nd}, {nx}, {ny}) modes: {tr.get_num_modes()}')

    num_modes = tr.get_num_modes()
    if num_modes == 0:
        raise Exception(f'ERROR: There are no modes for this transformation')
    
    for mode in range(num_modes):
        
        # -------------------------------------------------
        ws = wb.create_sheet(str(mode))
        wb.active = ws
        ws.cell(1, 1).value = n
        ws.cell(2, 1).value = T
        ws.cell(3, 1).value = nx
        ws.cell(3, 2).value = ny

        results = []
        outs = {}

        row = 5
        print(f'num forms: {tr.get_num_forms(mode)}')
        for form in range(tr.get_num_forms(mode)):
            txt = f'mode: {mode+1}/{tr.get_num_modes()}, form: {form+1}/{tr.get_num_forms(mode)}'
            print(txt)

            ws.cell(row, 1).value = nx
            ws.cell(row, 2).value = ny
            row += 1

            output_form = tr.get_output_form(mode, form)
            ones = count_ones(output_form.values())
            ws.cell(row, 1).value = txt
            row += 1

            print(ny, mode, ones)
            print(tr.get_plug_input())
            print(output_form)

            ws.cell(row, 1).value = str(output_form)
            row += 1

            ws.cell(row=row, column=1).value = '-'
            row += 1
            
            row = render(ws, row, dim, T, n, tr, mode, form, results, outs)

        # -------------------------------------------------
        ws = wb.create_sheet(f'mode {mode}')
        row = 1

        count = 0
        num_outs = len(outs)

        for index in outs.keys():
            result = results[index]

            ws.cell(row=row, column=1).value = f'type: {count+1}/{num_outs}'
            count += 1
            row += 1

            num_forms = len(outs[index])
            ws.cell(row=row, column=1).value = f'num_forms: {num_forms}'
            row += 1

            for output in outs[index]:
                ws.cell(row=row, column=1).value = str(output)
                row += 1

            ws.cell(row=row, column=1).value = '-'
            row += 1

            if dim == 1:
                pos = 0
                for x in range(T+1):
                    ws.cell(row=row, column=x+1).value = result[pos]
                    pos += 1
                row += 3

            elif dim == 2:
                pos = 0
                for _ in range(T+1):
                    for x in range(T+1):
                        ws.cell(row=row, column=x+1).value = result[pos]
                        pos += 1
                    row += 1
                row += 3

        # -------------------------------------------------
        wb.create_sheet(f'grid {mode}')

    print(f'saving: {fname}')
    wb.save(fname)


if __name__ == '__main__':
    freeze_support()
    if not (5 < len(sys.argv) < 8):
        print(f'syntax: test_transform.py <dim> <n> <T> <nd> <nx> [<ny> [<nz>]]')
        exit()
    dim = int(sys.argv[1])
    n = int(sys.argv[2])
    T = int(sys.argv[3])
    nd = int(sys.argv[4])

    nx = sys.argv[5]
    nx_begin = int(nx) if '-' not in nx else int(nx.split('-')[0])
    nx_end   = int(nx) if '-' not in nx else int(nx.split('-')[1])
    ny = '0'
    ny_begin = int(ny)
    ny_end   = int(ny)
    nz = '0'
    nz_begin = int(nz)
    nz_end   = int(nz)
    
    if len(sys.argv) > 6:
        ny = sys.argv[6]
        ny_begin = int(ny) if '-' not in ny else int(ny.split('-')[0])
        ny_end   = int(ny) if '-' not in ny else int(ny.split('-')[1])
        if len(sys.argv) > 7:
            nz = sys.argv[7]
            nz_begin = int(nz) if '-' not in nz else int(nz.split('-')[0])
            nz_end   = int(nz) if '-' not in nz else int(nz.split('-')[1])

    begin = time()

    for nx in range(nx_begin, nx_end+1):
        for ny in range(ny_begin, ny_end+1):
            for nz in range(nz_begin, nz_end+1):
                folder = f'{dim}d/{n}.{T}.{nd}/{nx:02d}'
                if not os.path.exists(f'{folder}'):
                    os.makedirs(f'{folder}')
                fname = f'{folder}/{dim}d.{n}.{T}.{nd}.{nx:02d}.{ny:02d}.{nz:02d}.xlsx'
                main3(fname, dim, n, T, nd, nx, ny, nz)

    end = time()
    duration = int(end - begin)
    secs = duration % 60
    min = (duration // 60) % 60
    hours = (duration // 3600) % 60
    print(f'{hours:02d}:{min:02d}:{secs:02d}')
