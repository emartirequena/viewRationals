import sys
import os
from openpyxl import Workbook

from utils import getDivisorsOfPeriod

root = r'C:\Users\emart\OneDrive\Documentos\Enrique\ArticuloRacionales\hojas'


def main(maxT: int, base: int, name: str):
    output = []
    for T in range(2, maxT + 1, 2):
        n = int(base**(T/2) + 1)
        a = (T*2) + 2
        print(T, n)

        nums = getDivisorsOfPeriod(n, T, base)

        for num in nums:
            output.append((num, T, a))
    
    print('Sorting...')
    output.sort(key=lambda x: x[0])

    print('Dump file...')
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet('main')
    wb.active = ws
    ws.cell(1, 4).value = 'Z'
    ws.cell(1, 5).value = 'mass num'
    ws.cell(1, 7).value = 'Divisor'
    ws.cell(1, 8).value = 'T'
    ws.cell(1, 9).value = '1/a'
    ws.cell(1, 10).value = 'Divisor*T'
    ws.cell(1, 11).value = 'log(Divisor*T)'
    ws.cell(1, 12).value = 'log(Divisor)'
    ws.cell(1, 13).value = 'Ajuste'

    row = 2
    for out in output:
        ws.cell(row, 7).value = out[0]
        ws.cell(row, 8).value = out[1]
        ws.cell(row, 9).value = out[2]
        row += 1

    fname = os.path.join(root, name)
    if not os.path.exists(root):
        os.makedirs(root)
    wb.save(fname)


if __name__ == '__main__':
    if len(sys.argv) != 4:
        print(f'Syntax: python test_mass.py <maxT> <base> <name>')
        exit()

    main(int(sys.argv[1]), int(sys.argv[2]), sys.argv[3])


