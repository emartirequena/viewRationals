import json
import os
from config import Config
from openpyxl import Workbook
from random import randint
from multiprocessing import Pool, freeze_support, get_context
from multiprocessing.managers import BaseManager
from utils import getPeriod, divisors
from timing import timing

def main():
    config = Config()
    path = config.get('files_path')
    file_paths = [
        os.path.join(path, '1D', 'P06', '1D_N63_P06_F3^2_7.json'),
        os.path.join(path, '1D', 'P08', '1D_N255_P08_F3_5_17.json'),
        os.path.join(path, '1D', 'P10', '1D_N1023_P10_F3_11_31.json'),
        os.path.join(path, '1D', 'P12', '1D_N4095_P12_F3^2_5_7_13.json'),
        os.path.join(path, '1D', 'P14', '1D_N16383_P14_F3_43_127.json'),
        os.path.join(path, '1D', 'P16', '1D_N65535_P16_F3_5_17_257.json'),
        os.path.join(path, '1D', 'P18', '1D_N262143_P18_F3^3_7_19_73.json')
    ]

    period = 6

    wb = Workbook()
    wb.remove(wb.active)
    for file_path in file_paths:
        print(f'Period: {period:02d}')

        with open(file_path, 'rt') as fp:
            content = json.load(fp)

        ws = wb.create_sheet(f'P{period:02d}')

        ws.cell(row=1, column=1, value='pos')
        ws.cell(row=1, column=2, value='time')
        ws.cell(row=1, column=3, value='count')

        last = content['spaces'][str(period)]
        row = 2
        for item in last:
            pos = item['pos'][0]
            time = item['time']
            count = item['count']
            ws.cell(row=row, column=1, value=pos)
            ws.cell(row=row, column=2, value=time)
            ws.cell(row=row, column=3, value=count)
            row += 1
        
        period += 2

    wb.save(os.path.join(path, '1D', 'times_new.xlsx'))

def main2():
    upper = 10000
    divupper = upper / 2.0

    print('   n  p1d  p2d  p3d')
    print('---- ---- ---- ----')
    num_p1d = 0
    num_p2d = 0
    num_p3d = 0
    for n in range(1, upper+1, 2):
        p1d = getPeriod(n, 2)
        p2d = getPeriod(n, 4)
        p3d = getPeriod(n, 8)
        # print(f'{n:4d} {p1d:4d} {p2d:4d} {p3d:4d}')

        if p1d % 2 == 1:
            num_p1d += 1
        if p2d % 2 == 1:
            num_p2d += 1
        if p3d % 2 == 1:
            num_p3d += 1
        
    print('')
    print(f'Odd period 1d: {100.0 * num_p1d / divupper:5.2f}%')
    print(f'Odd period 2d: {100.0 * num_p2d / divupper:5.2f}%')
    print(f'Odd period 3d: {100.0 * num_p3d / divupper:5.2f}%')


def main3():
    divisors_count = 0
    centrals_count = 0

    print('  T    div  count central  count   ratio')
    print('--- ------ ------ ------- ------ -------')

    total_set = set()

    for T in range(2, 46, 2):
        divisors_set = set(divisors(8**T-1))
        divisors_set.difference_update(total_set)
        divisors_len = len(divisors_set)
        divisors_count += divisors_len

        centrals_set = set(divisors(8**(T//2)+1))
        centrals_set.difference_update(total_set)
        centrals_len = len(centrals_set)
        centrals_count += centrals_len

        total_set.update(divisors_set)

        ratio = 100.0 * float(centrals_count) / float(divisors_count)

        print(f'{T:3d} {divisors_len:6d} {divisors_count:6d} {centrals_len:7d} {centrals_count:6d}  {ratio:5,.2f}%')


@timing
def main4():
    def cmp_key(a, b):
        if a < b:
            return -1
        elif a > b:
            return 1
        return 0

    def find_insert(plist, init, end, key, cmp):
        print(key, init, end)
        if len(plist) == 0:
            plist.append(key)
            return 0
        # elif end == init:
        #     plist.insert(init, key)
        #     return 
        elif end == init + 1:
            res = cmp(plist[init], key)
            if res < 0:
                plist.insert(end, key)
                return end
            elif res > 0:
                plist.insert(init, key)
                return init
        else:
            position = (init + end) // 2
            res = cmp(plist[position], key)
            if res == 0:
                # plist.insert(position, key)
                return position
            elif res > 0:
                return find_insert(plist, init, position, key, cmp)
            else:
                return find_insert(plist, position, end, key, cmp)

    plist = [0, 1, 1, 1, 1, 2, 2, 2, 2, 2, 2, 2, 2, 3, 3, 3, 3, 3]
    n = 2
    position = find_insert(plist, 0, len(plist), n, cmp_key)
    print(n, position, plist)


class MyClass:
    def __init__(self, val):
        self.val = val

    def set(self, val):
        print(f'set({val})')
        self.val = val

    def get(self):
        return self.val
    
class MyManager(BaseManager):
    pass

MyManager.register("MyClass", MyClass)

def modify(args):
    myclass, v = args
    print('----------')
    print(myclass)
    print(f'prev get() = {myclass.get()}')
    myclass.set(v)
    print(f'last get() = {myclass.get()}')
    print('----------')

def main5():
    with MyManager() as manager:
        myclass = manager.MyClass(-1)
        pool = Pool(1)
        pool.imap(func=modify, iterable=[(myclass, v) for v in range(3)])
        pool.close()
        pool.join()

        print("Last value stored:", myclass.get())

if __name__ == '__main__':
    freeze_support()
    main5()
